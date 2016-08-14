import openpyxl

from openpyxl.cell import get_column_letter
from openpyxl.cell import column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))

print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet3')
print(sheet)

print(type(sheet))

print(sheet.title)

another_sheet = wb.active

print(another_sheet)

sheet = wb.get_sheet_by_name('Sheet1')

print(sheet['A1'])

print(sheet['A1'].value)

cell = sheet['B1']

print(cell.value)

print('Row' + str(cell.row) + ', Column ' + cell.column + ' is ' + cell.value)
print('Cell ' + cell.coordinate + ' is ' + cell.value)

print(sheet['C1'].value)

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))
