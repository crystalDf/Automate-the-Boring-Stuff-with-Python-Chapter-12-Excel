import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'

wb.save('merged.xlsx')

sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')

wb.save('merged.xlsx')
